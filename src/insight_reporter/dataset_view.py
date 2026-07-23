"""Source-aware, format-independent access to retained tabular datasets."""

import csv
import hashlib
import io
import json
import math
import re
import unicodedata
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path, PurePosixPath
from typing import Any, Protocol

SUPPORTED_FORMATS = frozenset({"csv", "json", "xlsx"})
_ALLOWED_CONTROL_CHARACTERS = {"\t", "\n", "\r"}
_XLSX_REQUIRED_MEMBERS = frozenset({"[Content_Types].xml", "xl/workbook.xml"})
_MAX_XLSX_ARCHIVE_MEMBERS = 2_000
_MAX_XLSX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
_MAX_XLSX_COMPRESSION_RATIO = 100


class DatasetViewError(ValueError):
    """Raised when retained source data cannot produce a safe dataset view."""


@dataclass(frozen=True)
class SourceManifest:
    """Traceable immutable metadata for one retained source table."""

    source_id: str
    format: str
    internal_filename: str
    sha256: str
    row_count: int
    column_count: int
    table_name: str | None = None

    def to_dict(self) -> dict[str, object]:
        return {
            "source_id": self.source_id,
            "format": self.format,
            "internal_filename": self.internal_filename,
            "sha256": self.sha256,
            "row_count": self.row_count,
            "column_count": self.column_count,
            "table_name": self.table_name,
        }


@dataclass(frozen=True)
class ColumnReference:
    """Unambiguous reference to a column within a project source."""

    source_id: str
    column: str

    def to_dict(self) -> dict[str, str]:
        return {"source_id": self.source_id, "column": self.column}


@dataclass(frozen=True)
class DatasetRow:
    number: int
    values: dict[str, str]


class DatasetView(Protocol):
    """Interface consumed by deterministic analysis and formula evaluation."""

    @property
    def sources(self) -> tuple[SourceManifest, ...]: ...

    @property
    def headers(self) -> tuple[str, ...]: ...

    def iter_rows(self) -> tuple[DatasetRow, ...]: ...


@dataclass(frozen=True)
class _TabularDatasetView:
    source: SourceManifest
    _headers: tuple[str, ...]
    _rows: tuple[DatasetRow, ...]

    @property
    def sources(self) -> tuple[SourceManifest, ...]:
        return (self.source,)

    @property
    def headers(self) -> tuple[str, ...]:
        return self._headers

    def iter_rows(self) -> tuple[DatasetRow, ...]:
        return self._rows


@dataclass(frozen=True)
class CsvDatasetView(_TabularDatasetView):
    """Validated view over one retained UTF-8 CSV."""

    @classmethod
    def from_path(
        cls,
        path: Path,
        *,
        max_rows: int | None = None,
        max_columns: int | None = None,
    ) -> "CsvDatasetView":
        raw_bytes = _read_bytes(path)
        try:
            text = raw_bytes.decode("utf-8-sig", errors="strict")
        except UnicodeDecodeError as error:
            raise DatasetViewError(
                "File must be UTF-8 CSV text; binary or non-UTF-8 content was rejected."
            ) from error
        _validate_safe_text(text)
        if not text.strip():
            raise DatasetViewError("CSV is empty.")

        reader = csv.reader(io.StringIO(text, newline=""), strict=True)
        try:
            raw_headers = next(reader)
        except StopIteration as error:
            raise DatasetViewError("CSV is empty.") from error
        except csv.Error as error:
            raise DatasetViewError(f"Malformed CSV near line 1: {error}.") from error
        headers = _validate_headers(
            raw_headers,
            source_label="CSV",
            max_columns=max_columns,
        )

        rows: list[DatasetRow] = []
        try:
            for raw_row in reader:
                if not raw_row:
                    continue
                if len(raw_row) != len(headers):
                    raise DatasetViewError(
                        f"Malformed CSV row near line {reader.line_num}: expected "
                        f"{len(headers)} columns but found {len(raw_row)}."
                    )
                _validate_row_limit(len(rows) + 1, max_rows=max_rows, label="CSV")
                for value in raw_row:
                    _validate_safe_text(value)
                rows.append(
                    DatasetRow(
                        reader.line_num,
                        dict(zip(headers, raw_row, strict=True)),
                    )
                )
        except csv.Error as error:
            raise DatasetViewError(
                f"Malformed CSV near line {reader.line_num}: {error}."
            ) from error
        _require_rows(rows, label="CSV")
        return cls(
            source=_source_manifest(
                path,
                source_format="csv",
                row_count=len(rows),
                column_count=len(headers),
            ),
            _headers=headers,
            _rows=tuple(rows),
        )


@dataclass(frozen=True)
class JsonDatasetView(_TabularDatasetView):
    """Validated view over a top-level JSON array of flat objects."""

    @classmethod
    def from_path(
        cls,
        path: Path,
        *,
        max_rows: int | None = None,
        max_columns: int | None = None,
    ) -> "JsonDatasetView":
        raw_bytes = _read_bytes(path)
        try:
            text = raw_bytes.decode("utf-8-sig", errors="strict")
        except UnicodeDecodeError as error:
            raise DatasetViewError(
                "JSON must be UTF-8 text; binary or non-UTF-8 content was rejected."
            ) from error
        _validate_safe_text(text)
        if not text.strip():
            raise DatasetViewError("JSON is empty.")
        try:
            payload = json.loads(
                text,
                object_pairs_hook=_unique_json_object,
                parse_constant=_reject_json_constant,
            )
        except DatasetViewError:
            raise
        except (json.JSONDecodeError, ValueError) as error:
            if not isinstance(error, json.JSONDecodeError):
                raise DatasetViewError("JSON contains an invalid numeric value.") from error
            raise DatasetViewError(
                f"Malformed JSON near line {error.lineno}, column {error.colno}."
            ) from error
        if not isinstance(payload, list):
            raise DatasetViewError(
                "JSON must contain a top-level array of flat record objects."
            )
        if not payload:
            raise DatasetViewError("JSON must contain at least one record.")
        _validate_row_limit(len(payload), max_rows=max_rows, label="JSON")

        normalized_records: list[dict[str, str]] = []
        headers: list[str] = []
        normalized_header_names: dict[str, str] = {}
        for index, item in enumerate(payload, start=1):
            if not isinstance(item, dict):
                raise DatasetViewError(
                    f"JSON record {index} must be an object, not a nested or scalar value."
                )
            record: dict[str, str] = {}
            for raw_key, raw_value in item.items():
                clean_key = raw_key.strip()
                if not clean_key:
                    raise DatasetViewError(
                        f"JSON record {index} contains an empty column name."
                    )
                normalized_key = _normalized_header(clean_key)
                existing = normalized_header_names.get(normalized_key)
                if existing is not None and existing != clean_key:
                    raise DatasetViewError(
                        "JSON contains duplicate column names after normalization."
                    )
                if existing is None:
                    normalized_header_names[normalized_key] = clean_key
                    headers.append(clean_key)
                    if max_columns is not None and len(headers) > max_columns:
                        raise DatasetViewError(
                            f"JSON has more than the maximum of {max_columns} columns."
                        )
                record[clean_key] = _json_scalar(raw_value, record_number=index)
            normalized_records.append(record)
        safe_headers = _validate_headers(
            headers,
            source_label="JSON",
            max_columns=max_columns,
        )
        rows = tuple(
            DatasetRow(
                index,
                {header: record.get(header, "") for header in safe_headers},
            )
            for index, record in enumerate(normalized_records, start=1)
        )
        return cls(
            source=_source_manifest(
                path,
                source_format="json",
                row_count=len(rows),
                column_count=len(safe_headers),
            ),
            _headers=safe_headers,
            _rows=rows,
        )


@dataclass(frozen=True)
class XlsxDatasetView(_TabularDatasetView):
    """Validated view over one explicitly selected XLSX worksheet."""

    @classmethod
    def from_path(
        cls,
        path: Path,
        *,
        table_name: str,
        max_rows: int | None = None,
        max_columns: int | None = None,
    ) -> "XlsxDatasetView":
        _validate_xlsx_archive(path)
        workbook = _load_xlsx_workbook(path)
        try:
            visible_sheets = _visible_sheet_names(workbook)
            if table_name not in visible_sheets:
                raise DatasetViewError("Selected Excel worksheet is unavailable.")
            worksheet = workbook[table_name]
            if max_columns is not None and worksheet.max_column > max_columns:
                raise DatasetViewError(
                    f"Excel worksheet has {worksheet.max_column} columns; "
                    f"the maximum is {max_columns}."
                )
            if max_rows is not None and max(0, worksheet.max_row - 1) > max_rows:
                raise DatasetViewError(
                    f"Excel worksheet has more than the maximum of {max_rows} data rows."
                )
            iterator = worksheet.iter_rows()
            try:
                header_cells = next(iterator)
            except StopIteration as error:
                raise DatasetViewError("Excel worksheet is empty.") from error
            raw_headers: list[str] = []
            for cell in header_cells:
                if cell.data_type == "f":
                    raise DatasetViewError(
                        "Excel header formulas are not supported; use literal column names."
                    )
                if not isinstance(cell.value, str):
                    raise DatasetViewError(
                        "Excel worksheet headers must be non-empty text values in the first row."
                    )
                raw_headers.append(cell.value)
            headers = _validate_headers(
                raw_headers,
                source_label="Excel worksheet",
                max_columns=max_columns,
            )

            rows: list[DatasetRow] = []
            for sheet_row_number, cells in enumerate(iterator, start=2):
                values = tuple(_xlsx_cell_value(cell) for cell in cells[: len(headers)])
                if all(value == "" for value in values):
                    continue
                _validate_row_limit(
                    len(rows) + 1,
                    max_rows=max_rows,
                    label="Excel worksheet",
                )
                rows.append(
                    DatasetRow(
                        sheet_row_number,
                        dict(zip(headers, values, strict=True)),
                    )
                )
            _require_rows(rows, label="Excel worksheet")
        finally:
            workbook.close()
        return cls(
            source=_source_manifest(
                path,
                source_format="xlsx",
                row_count=len(rows),
                column_count=len(headers),
                table_name=table_name,
            ),
            _headers=headers,
            _rows=tuple(rows),
        )


def load_dataset_view(
    path: Path,
    *,
    table_name: str | None = None,
    max_rows: int | None = None,
    max_columns: int | None = None,
) -> DatasetView:
    """Load a retained dataset according to its server-controlled extension."""

    source_format = path.suffix.removeprefix(".").casefold()
    if source_format == "csv":
        return CsvDatasetView.from_path(
            path, max_rows=max_rows, max_columns=max_columns
        )
    if source_format == "json":
        return JsonDatasetView.from_path(
            path, max_rows=max_rows, max_columns=max_columns
        )
    if source_format == "xlsx":
        if not table_name:
            raise DatasetViewError("Select an Excel worksheet before profiling.")
        return XlsxDatasetView.from_path(
            path,
            table_name=table_name,
            max_rows=max_rows,
            max_columns=max_columns,
        )
    raise DatasetViewError("Retained dataset format is unsupported.")


def detect_dataset_format(raw_bytes: bytes) -> str:
    """Detect supported content without trusting filename or MIME metadata."""

    if raw_bytes.startswith(b"PK"):
        try:
            with zipfile.ZipFile(io.BytesIO(raw_bytes)) as archive:
                names = set(archive.namelist())
        except (OSError, zipfile.BadZipFile) as error:
            raise DatasetViewError("Binary ZIP content is not a valid XLSX workbook.") from error
        if _XLSX_REQUIRED_MEMBERS.issubset(names):
            return "xlsx"
        raise DatasetViewError("ZIP archives are not supported dataset inputs.")
    try:
        text = raw_bytes.decode("utf-8-sig", errors="strict")
    except UnicodeDecodeError as error:
        raise DatasetViewError(
            "Binary or non-UTF-8 content is not a supported dataset."
        ) from error
    _validate_safe_text(text)
    first = text.lstrip()[:1]
    return "json" if first in {"[", "{"} else "csv"


def discover_xlsx_tables(path: Path) -> tuple[str, ...]:
    """Return visible worksheet names after validating the workbook container."""

    _validate_xlsx_archive(path)
    workbook = _load_xlsx_workbook(path)
    try:
        names = _visible_sheet_names(workbook)
    finally:
        workbook.close()
    if not names:
        raise DatasetViewError("Excel workbook has no visible worksheets.")
    return names


def source_id_from_hash(sha256: str, table_name: str | None = None) -> str:
    """Return a stable source-table ID without exposing an entire content hash."""

    if re.fullmatch(r"[0-9a-f]{64}", sha256) is None:
        raise DatasetViewError("Source hash is invalid.")
    if table_name is None:
        return f"SRC-{sha256[:12].upper()}"
    identity = f"{sha256}\0{table_name}"
    digest = hashlib.sha256(identity.encode("utf-8")).hexdigest()
    return f"SRC-{digest[:12].upper()}"


def load_source_manifest(payload: object) -> SourceManifest:
    """Load old CSV and current multi-format persisted source manifests."""

    old_expected = {
        "source_id",
        "format",
        "internal_filename",
        "sha256",
        "row_count",
        "column_count",
    }
    current_expected = old_expected | {"table_name"}
    if not isinstance(payload, dict) or set(payload) not in {
        frozenset(old_expected),
        frozenset(current_expected),
    }:
        raise DatasetViewError("Saved source manifest has an invalid shape.")
    source_id = payload.get("source_id")
    source_format = payload.get("format")
    internal_filename = payload.get("internal_filename")
    sha256 = payload.get("sha256")
    row_count = payload.get("row_count")
    column_count = payload.get("column_count")
    table_name = payload.get("table_name")
    if not all(
        isinstance(value, str)
        for value in (source_id, source_format, internal_filename, sha256)
    ):
        raise DatasetViewError("Saved source manifest contains invalid text.")
    if source_format not in SUPPORTED_FORMATS:
        raise DatasetViewError("Saved source format is not supported.")
    if re.fullmatch(
        rf"[0-9a-f]{{32}}\.{re.escape(source_format)}", internal_filename
    ) is None:
        raise DatasetViewError("Saved source filename is invalid.")
    if source_format == "xlsx":
        if (
            not isinstance(table_name, str)
            or not table_name
            or len(table_name) > 31
            or not table_name.isprintable()
        ):
            raise DatasetViewError("Saved Excel worksheet name is invalid.")
    elif table_name is not None:
        raise DatasetViewError("Saved non-Excel source contains an invalid table name.")
    if source_id != source_id_from_hash(sha256, table_name):
        raise DatasetViewError("Saved source ID does not match its source table.")
    if (
        isinstance(row_count, bool)
        or not isinstance(row_count, int)
        or row_count < 1
        or isinstance(column_count, bool)
        or not isinstance(column_count, int)
        or column_count < 1
    ):
        raise DatasetViewError("Saved source dimensions are invalid.")
    return SourceManifest(
        source_id=source_id,
        format=source_format,
        internal_filename=internal_filename,
        sha256=sha256,
        row_count=row_count,
        column_count=column_count,
        table_name=table_name,
    )


def load_column_reference(
    payload: object, *, sources: tuple[SourceManifest, ...]
) -> ColumnReference:
    """Load a source-qualified column reference."""

    if not isinstance(payload, dict) or set(payload) != {"source_id", "column"}:
        raise DatasetViewError("Saved column reference has an invalid shape.")
    source_id = payload.get("source_id")
    column = payload.get("column")
    if not isinstance(source_id, str) or not isinstance(column, str) or not column:
        raise DatasetViewError("Saved column reference contains invalid text.")
    if source_id not in {source.source_id for source in sources}:
        raise DatasetViewError("Saved column reference uses an unknown source.")
    return ColumnReference(source_id=source_id, column=column)


def _read_bytes(path: Path) -> bytes:
    try:
        return path.read_bytes()
    except OSError as error:
        raise DatasetViewError("Retained dataset cannot be read safely.") from error


def _source_manifest(
    path: Path,
    *,
    source_format: str,
    row_count: int,
    column_count: int,
    table_name: str | None = None,
) -> SourceManifest:
    raw_bytes = _read_bytes(path)
    digest = hashlib.sha256(raw_bytes).hexdigest()
    return SourceManifest(
        source_id=source_id_from_hash(digest, table_name),
        format=source_format,
        internal_filename=path.name,
        sha256=digest,
        row_count=row_count,
        column_count=column_count,
        table_name=table_name,
    )


def _validate_headers(
    raw_headers: list[str] | tuple[str, ...],
    *,
    source_label: str,
    max_columns: int | None,
) -> tuple[str, ...]:
    headers = tuple(header.strip() for header in raw_headers)
    if not headers or any(not header for header in headers):
        raise DatasetViewError(f"{source_label} header contains an empty column name.")
    for header in headers:
        _validate_safe_text(header)
    if max_columns is not None and len(headers) > max_columns:
        raise DatasetViewError(
            f"{source_label} has {len(headers)} columns; the maximum is {max_columns}."
        )
    normalized = [_normalized_header(header) for header in headers]
    if len(normalized) != len(set(normalized)):
        raise DatasetViewError(f"{source_label} contains duplicate column names.")
    return headers


def _normalized_header(value: str) -> str:
    return unicodedata.normalize("NFKC", value).casefold()


def _validate_safe_text(text: str) -> None:
    if any(
        unicodedata.category(character) == "Cc"
        and character not in _ALLOWED_CONTROL_CHARACTERS
        for character in text
    ):
        raise DatasetViewError("Binary or unsafe control characters were detected.")


def _validate_row_limit(
    row_count: int, *, max_rows: int | None, label: str
) -> None:
    if max_rows is not None and row_count > max_rows:
        raise DatasetViewError(
            f"{label} has more than the maximum of {max_rows} data rows."
        )


def _require_rows(rows: list[DatasetRow], *, label: str) -> None:
    if not rows:
        raise DatasetViewError(f"{label} must contain at least one data row.")


def _unique_json_object(pairs: list[tuple[str, object]]) -> dict[str, object]:
    result: dict[str, object] = {}
    for key, value in pairs:
        if key in result:
            raise DatasetViewError(f"JSON object contains duplicate key: {key}.")
        result[key] = value
    return result


def _reject_json_constant(value: str) -> object:
    raise DatasetViewError(f"JSON contains unsupported numeric constant: {value}.")


def _json_scalar(value: object, *, record_number: int) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, str):
        _validate_safe_text(value)
        return value
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            raise DatasetViewError(
                f"JSON record {record_number} contains a non-finite number."
            )
        return repr(value)
    raise DatasetViewError(
        f"JSON record {record_number} contains a nested array or object value."
    )


def _validate_xlsx_archive(path: Path) -> None:
    try:
        with zipfile.ZipFile(path) as archive:
            infos = archive.infolist()
            names = {info.filename for info in infos}
            if not _XLSX_REQUIRED_MEMBERS.issubset(names):
                raise DatasetViewError("File is not a valid XLSX workbook.")
            if len(infos) > _MAX_XLSX_ARCHIVE_MEMBERS:
                raise DatasetViewError("Excel workbook contains too many archive members.")
            total_uncompressed = 0
            for info in infos:
                member_path = PurePosixPath(info.filename)
                if (
                    member_path.is_absolute()
                    or ".." in member_path.parts
                    or info.flag_bits & 0x1
                ):
                    raise DatasetViewError("Excel workbook archive structure is unsafe.")
                total_uncompressed += info.file_size
                if total_uncompressed > _MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise DatasetViewError("Excel workbook expands beyond the safe limit.")
                if (
                    info.file_size > 0
                    and info.compress_size > 0
                    and info.file_size / info.compress_size
                    > _MAX_XLSX_COMPRESSION_RATIO
                ):
                    raise DatasetViewError(
                        "Excel workbook contains an unsafe compression ratio."
                    )
            if any(
                name.casefold().endswith("vbaproject.bin")
                or name.casefold().startswith("xl/externallinks/")
                for name in names
            ):
                raise DatasetViewError(
                    "Macro-enabled or externally linked Excel workbooks are not supported."
                )
            for info in infos:
                if info.filename.casefold().endswith(".xml"):
                    with archive.open(info) as member:
                        prefix = member.read(1024 * 1024).upper()
                    if b"<!DOCTYPE" in prefix or b"<!ENTITY" in prefix:
                        raise DatasetViewError(
                            "Excel workbook contains unsafe XML declarations."
                        )
    except DatasetViewError:
        raise
    except (OSError, zipfile.BadZipFile, RuntimeError) as error:
        raise DatasetViewError("Excel workbook archive is unreadable.") from error


def _load_xlsx_workbook(path: Path):  # type: ignore[no-untyped-def]
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise DatasetViewError(
            "Excel support is unavailable because the openpyxl dependency is not installed."
        ) from error
    try:
        return load_workbook(
            path,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except Exception as error:
        raise DatasetViewError("Excel workbook could not be read safely.") from error


def _visible_sheet_names(workbook: Any) -> tuple[str, ...]:
    return tuple(
        worksheet.title
        for worksheet in workbook.worksheets
        if worksheet.sheet_state == "visible"
    )


def _xlsx_cell_value(cell: Any) -> str:
    if cell.data_type == "f":
        raise DatasetViewError(
            "Excel formula cells are not supported; replace formulas with values."
        )
    if cell.data_type == "e":
        raise DatasetViewError("Excel error cells are not supported.")
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime | date | time):
        return value.isoformat()
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            raise DatasetViewError("Excel worksheet contains a non-finite number.")
        return repr(value)
    if isinstance(value, str):
        _validate_safe_text(value)
        return value
    raise DatasetViewError("Excel worksheet contains an unsupported cell value.")
